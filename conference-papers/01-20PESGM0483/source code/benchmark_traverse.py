# !/usr/bin/python
#
# Copyright 2019, Southeast University, Liu Pengxiang
#
# IEEE PES-GM 2019 Paper
# A novel acceleration of N-1 contingency screening for distribution system


import math
import time
import numpy as np
import openpyxl as pyxl
import matplotlib.pyplot as plt

from gurobipy import *
from datetime import datetime


# This class creates the parameter class
#
class Parameter(object):

    # Initialization
    def __init__(self, filename):
        # System Data
        Data = self.ReadData(filename)
        # Base value
        self.Base_V = 12.66  # voltage: 12.66 kV
        self.Base_S = 10.00  # power:   10.00 MVA
        self.Base_Z = self.Base_V ** 2 / self.Base_S  # impedance
        self.Base_I = self.Base_S / self.Base_V / np.sqrt(3)  # current
        # Line
        self.Line   = Data[0]
        self.N_line = len(self.Line) # number of line
        # Bus
        self.Bus    = Data[1]
        self.N_bus  = len(self.Bus)  # number of bus
        # Sub
        self.Sub    = Data[2]
        self.N_sub  = len(self.Sub)  # number of substation
        # Gen
        self.Gen    = Data[3]
        self.N_gen  = len(self.Gen)  # number of renewables
        self.Factor = 0.31756  # power factor (rad)
        # Day
        self.Day    = Data[4]
        self.N_time = len(self.Day)  # number of hours
        # Cost
        self.Cost_sub = 83   # cost of power purchasing
        self.Cost_pen = 200  # cost of load shedding
        self.Cost_los = 25   # cost of power loss
        # Other
        self.Big_M = 1e2  # a sufficient large number
        self.V_min = (0.95 * self.Base_V) ** 2
        self.V_max = (1.05 * self.Base_V) ** 2
        # Bus-Line Information
        self.Line_head = [[] for i in range(self.N_bus)]
        self.Line_tail = [[] for i in range(self.N_bus)]
        for i in range(self.N_line):
            head = self.Line[i][1]
            tail = self.Line[i][2]
            self.Line_head[int(round(head))].append(i)
            self.Line_tail[int(round(tail))].append(i)

    # This function inputs data from Excel files
    def ReadData(self, filename):
        data = []
        book = pyxl.load_workbook(filename)
        # Data preprocessing
        for i in book.sheetnames:  # sheet number
            sheet = book[i]
            n_row = sheet.max_row     # number of rows
            n_col = sheet.max_column  # number of columns
            data.append(self.Matrix_slice(sheet, n_row, n_col))  # append data
        return data

    # This function slices the matrix based on data
    def Matrix_slice(self, sheet, n_row, n_col):
        matrix = []
        k = 0
        for i in range(n_row):
            if sheet['A' + str(i + 1)].data_type == 'n':  # if value is a number
                matrix.append([])
                for j in range(n_col):
                    pos = chr(64 + j + 1) + str(i + 1)  # the position in the sheet
                    val = sheet[pos].value
                    matrix[k].append(val)
                k = k + 1
        return np.array(matrix)


# This class creates the drawing class
#
class Drawing(object):

    # Initialization
    def __init__(self, Para):
        # Switch line coordinate
        self.switch_line = []
        self.switch_line.append(np.array([[ 7, 0], [ 7,-2], [3 ,-2], [ 3,-3]]))
        self.switch_line.append(np.array([[ 8, 0], [ 8, 2], [14, 2], [14, 0]]))
        self.switch_line.append(np.array([[11, 0], [11,-3], [ 4,-3]]))
        self.switch_line.append(np.array([[17, 0], [17, 3], [12, 3]]))
        self.switch_line.append(np.array([[ 8, 6], [ 8, 5], [ 8, 3]]))
        # Line coordinate
        self.coordinate = []
        for n in range(Para.N_line):
            if Para.Line[n,6] == 0:  # line
                bus_head = int(round(Para.Line[n,1]))
                bus_tail = int(round(Para.Line[n,2]))
                x0 = Para.Bus[bus_head, 3]
                y0 = Para.Bus[bus_head, 4]
                x1 = Para.Bus[bus_tail, 3]
                y1 = Para.Bus[bus_tail, 4]
                self.coordinate.append(np.array([[x0,y0], [x1, y1]]))
            else:
                switch_no = int(Para.Line[n,6] - 1)
                self.coordinate.append(self.switch_line[switch_no])
    
    # Figuring
    def Figuring(self, Para, Sol):
        self.Plot_Bus (Para)
        self.Plot_Line(Para, Sol)
        plt.axis('equal')
        plt.show()

    # Bus
    def Plot_Bus(self, Para):
        for n in range(Para.N_bus):
            plt.plot(Para.Bus[n,3], Para.Bus[n,4], 'b.')
            plt.text(Para.Bus[n,3] + 0.05, Para.Bus[n,4] + 0.10, '%s' % n)
    
    # Line
    def Plot_Line(self, Para, Sol):
        for n in range(Para.N_line):
            for m in range(np.size(self.coordinate[n], 0) - 1):
                x0 = self.coordinate[n][m, 0]
                y0 = self.coordinate[n][m, 1]
                x1 = self.coordinate[n][m + 1, 0]
                y1 = self.coordinate[n][m + 1, 1]
                if Sol.y_line[n,0] == 1:
                    plt.plot([x0, x1], [y0, y1], 'b-' )
                if Sol.y_line[n,0] == 0:
                    plt.plot([x0, x1], [y0, y1], 'r--')
    
    # Flow
    def Plot_Flow(self, Para, Sol):
        for n in range(Para.N_line):
            flow = Sol.P_line[n]
            x = 0.8 * self.coordinate[n][0, 0] + 0.2 * self.coordinate[n][1, 0]
            y = 0.8 * self.coordinate[n][0, 1] + 0.2 * self.coordinate[n][1, 1]
            plt.text(x, y - 0.35, '%.2f' % flow)


# This class restores the results of power flow
#
class Result_Flow(object):
    
    # Initialization
    def __init__(self, model, Para, sign, y_line, f_line, c_line, v_flow):
        self.sign   = sign.x
        self.y_line = self.value(y_line, 'int')
        self.f_line = self.value(f_line, 'int')
        self.c_line = self.value(c_line, 'int')
        self.v_flow = self.value(v_flow, 'float')
        self.V_bus  = self.v_flow[N_V_bus  : N_V_bus  + Para.N_bus ]
        self.I_line = self.v_flow[N_I_line : N_I_line + Para.N_line]
        self.P_line = self.v_flow[N_P_line : N_P_line + Para.N_line]
        self.Q_line = self.v_flow[N_Q_line : N_Q_line + Para.N_line]
        self.P_sub  = self.v_flow[N_P_sub  : N_P_sub  + Para.N_sub ]
        self.Q_sub  = self.v_flow[N_Q_sub  : N_Q_sub  + Para.N_sub ]
        self.S_gen  = self.v_flow[N_S_gen  : N_S_gen  + Para.N_gen ]
        self.C_gen  = self.v_flow[N_C_gen  : N_C_gen  + Para.N_gen ]
        self.P_cut  = self.v_flow[N_P_cut  : N_P_cut  + Para.N_bus ]
        self.Q_cut  = self.v_flow[N_Q_cut  : N_Q_cut  + Para.N_bus ]
    
    # Convert gurobi tuplelist to array
    def value(self,variable,string):
        # Get value
        key = variable.keys()
        val = variable.copy()
        for i in range(len(key)):
            val[key[i]] = variable[key[i]].x
        # Calculate dimention
        if isinstance(max(key),tuple):  # multi dimention
            dim = tuple([item + 1 for item in max(key)])
        if isinstance(max(key),int):    # one   dimention
            dim = tuple([int(len(key)),1])
        # Convert dictionary to numpy array
        arr = np.zeros(dim, dtype = string)
        for i in range(len(val)):
            arr[key[i]] = val[key[i]]
        return arr


# This function creates global index
#
def Global_Index(Para):

    '''---------------------------Fictitious power flow----------------------------'''
    # 1. Fictitious power flow
    global N_F_line, N_F_sub , N_F_load, N_F_gen , N_F_var
    # Initialization
    N_F_line = 0                       # flow of line
    N_F_load = N_F_line + Para.N_line  # flow of load demand
    N_F_sub  = N_F_load + Para.N_bus   # flow of substation
    N_F_gen  = N_F_sub  + Para.N_sub   # flow of DG
    N_F_var  = N_F_gen  + Para.N_gen   # Number of all fictitious vaariables

    '''------------------------------Real power flow-------------------------------'''
    # 2. Real power flow
    global N_V_bus, N_I_line, N_P_line, N_Q_line, N_P_sub, N_Q_sub
    global N_S_gen, N_C_gen , N_P_cut , N_Q_cut , N_N_var
    # Initialization
    N_V_bus  = 0                       # square of voltage amplitude
    N_I_line = N_V_bus  + Para.N_bus   # square of voltage phase angle
    N_P_line = N_I_line + Para.N_line  # power flow (active)
    N_Q_line = N_P_line + Para.N_line  # power flow (reactive)
    N_P_sub  = N_Q_line + Para.N_line  # power injection at substation
    N_Q_sub  = N_P_sub  + Para.N_sub   # power injection at substation
    N_S_gen  = N_Q_sub  + Para.N_sub   # renewables generation
    N_C_gen  = N_S_gen  + Para.N_gen   # renewables curtailment
    N_P_cut  = N_C_gen  + Para.N_gen   # Load shedding (active)
    N_Q_cut  = N_P_cut  + Para.N_bus   # Load shedding (reactive)
    N_N_var  = N_Q_cut  + Para.N_bus   # Number of all variables
    
    # Return
    return True


# This function defines a contingency screening model
#
def Contingency_Screening(Para, hour):
    
    '''------------------------------Initialization------------------------------'''
    # Import gurobi model
    model = Model()

    # Islanding sign
    sign = model.addVar(vtype = GRB.BINARY)
    # Topology variables
    y_line = model.addVars(Para.N_line, vtype = GRB.BINARY)  # reconfiguration
    c_line = model.addVars(Para.N_line, vtype = GRB.BINARY)  # contingency
    f_line = model.addVars(Para.N_line, vtype = GRB.BINARY)  # islanding mode
    # Operating variable
    f_flow = model.addVars(N_F_var, lb = -1e2)
    v_flow = model.addVars(N_N_var, lb = -1e2)
    # Set objective
    obj    = model.addVar ()

    '''-------------------------------Build Models-------------------------------'''
    # Build the model
    model = Contin_Model(model, Para, sign, y_line, f_line, c_line)
    model = Reconf_Model(model, Para, sign, y_line, f_line, f_flow)
    model = Operat_Model(model, Para, y_line, v_flow, hour, obj)

    # Objective
    model.setObjective(obj, GRB.MINIMIZE)

    '''-------------------------------Optimization-------------------------------'''
    # Additional constraints
    for n in range(Para.N_bus):
        model.addConstr(v_flow[N_P_cut + n] == 0)
        model.addConstr(v_flow[N_Q_cut + n] == 0)
    # Set parameters
    model.setParam("MIPGap", 0.1)
    model._s_line = np.ones((Para.N_line, 2))
    model.update()
    for n in range(Para.N_line):
        model_temp = model.copy()
        model_temp.update()
        var = model_temp.getVars()
        model_temp.addConstr(var[Para.N_line + 1 + n] == 1)
        model_temp.optimize()
        if model_temp.status == GRB.Status.OPTIMAL:
            model._s_line[n, 0] = 0
        if model_temp.status == GRB.Status.INFEASIBLE:
            continue
    sol = model._s_line
    return sol


# This function defines the contingency model
#
def Contin_Model(model, Para, sign, y_line, f_line, c_line):

    # Constraint

    # 0. Topology
    for n in range(Para.N_line):
        model.addConstr(c_line[n] - f_line[n] <= 1 - sign)
        model.addConstr(y_line[n] + c_line[n] <= 1)
        model.addConstr(y_line[n] + f_line[n] <= 1)
    
    # 1. N-1 contingency and islanding
    model.addConstr(c_line.sum('*') == 1)
    model.addConstr(f_line.sum('*') == sign)

    # Return
    return model


# This function defines the reconfiguration model
#
def Reconf_Model(model, Para, sign, y_line, f_line, f_flow):
    
    # Constraint

    # 0. Fictitious power flow
    for n in range(Para.N_line):
        model.addConstr(f_flow[N_F_line + n] >= -1e2 * (y_line[n] + f_line[n]))
        model.addConstr(f_flow[N_F_line + n] <=  1e2 * (y_line[n] + f_line[n]))
    for n in range(Para.N_sub):
        model.addConstr(f_flow[N_F_sub  + n] >=  0)
        model.addConstr(f_flow[N_F_sub  + n] <=  1e2)
    for n in range(Para.N_bus):
        model.addConstr(f_flow[N_F_load + n] ==  1)
    for n in range(Para.N_gen):
        model.addConstr(f_flow[N_F_gen  + n] == -1)

    # 1. Connectivity
    for n in range(Para.N_bus):
        # Bus-branch information
        line_head = Para.Line_head[n]
        line_tail = Para.Line_tail[n]
        # Formulate expression
        expr = LinExpr()
        expr = expr - f_flow[N_F_load + n]
        expr = expr - quicksum(f_flow[N_F_line + i] for i in line_head)
        expr = expr + quicksum(f_flow[N_F_line + i] for i in line_tail)
        if n in Para.Sub[:,1]:
            i = int(np.where(n == Para.Sub[:,1])[0])
            expr = expr + f_flow[N_F_sub + i]
        if n in Para.Gen[:,1]:
            i = int(np.where(n == Para.Gen[:,1])[0])
            expr = expr + f_flow[N_F_gen + i]
        model.addConstr(expr == 0)
    
    # 2. Radial topology
    model.addConstr(y_line.sum('*') == Para.N_bus - Para.N_sub - sign)
    
    # Return
    return model


# This function defines the operation model
#
def Operat_Model(model, Para, y_line, v_flow, hour, obj):
    
    # Objective
    opr = LinExpr()
    for n in range(Para.N_line):
        opr = opr + v_flow[N_I_line + n] * Para.Cost_los
    for n in range(Para.N_sub):  # power purchasing
        opr = opr + v_flow[N_P_sub  + n] * Para.Cost_sub
    for n in range(Para.N_bus):  # load shedding
        opr = opr + v_flow[N_P_cut  + n] * Para.Cost_pen
        opr = opr + v_flow[N_Q_cut  + n] * Para.Cost_pen
    for n in range(Para.N_gen):  # renewables
        opr = opr + v_flow[N_S_gen  + n] * Para.Gen[n,3]
    model.addConstr(obj == opr)
    
    # Constraint
    # 0. Nodal active power balance
    for n in range(Para.N_bus):
        # Bus-Line information
        line_head = Para.Line_head[n]
        line_tail = Para.Line_tail[n]
        # Formulate expression
        expr = LinExpr()
        expr = expr - quicksum(v_flow[N_P_line + i] for i in line_head)
        expr = expr + quicksum(v_flow[N_P_line + i] for i in line_tail)
        expr = expr + v_flow[N_P_cut + n]
        for i in line_tail:
            expr = expr - v_flow[N_I_line + i] * Para.Line[i,4]
        if n in Para.Sub[:,1]:  # active power input from substation
            i = int(np.where(n == Para.Sub[:,1])[0])
            expr = expr + v_flow[N_P_sub + i]
        if n in Para.Gen[:,1]:  # active power input from renewables
            i = int(np.where(n == Para.Gen[:,1])[0])
            expr = expr + v_flow[N_S_gen + i] * math.cos(Para.Factor)
        model.addConstr(expr == Para.Bus[n,1] * Para.Day[hour,1])
    
    # 1. Nodal reactive power balance
    for n in range(Para.N_bus):
        # Bus-Line information
        line_head = Para.Line_head[n]
        line_tail = Para.Line_tail[n]
        # Formulate expression
        expr = LinExpr()
        expr = expr - quicksum(v_flow[N_Q_line + i] for i in line_head)
        expr = expr + quicksum(v_flow[N_Q_line + i] for i in line_tail)
        expr = expr + v_flow[N_Q_cut + n]
        for i in line_tail:
            expr = expr - v_flow[N_I_line + i] * Para.Line[i,5]
        if n in Para.Sub[:,1]:  # active power input from substation
            i = int(np.where(n == Para.Sub[:,1])[0])
            expr = expr + v_flow[N_Q_sub + i]
        if n in Para.Gen[:,1]:  # active power input from renewables
            i = int(np.where(n == Para.Gen[:,1])[0])
            expr = expr + v_flow[N_S_gen + i] * math.sin(Para.Factor)
        model.addConstr(expr == Para.Bus[n,2] * Para.Day[hour,1])
    
    # 2. Branch flow equation
    for n in range(Para.N_line):
        bus_head = Para.Line[n,1]
        bus_tail = Para.Line[n,2]
        # Formulate expression
        expr = LinExpr()
        expr = expr + v_flow[N_V_bus + bus_head] - v_flow[N_V_bus + bus_tail]
        expr = expr - v_flow[N_P_line + n] * Para.Line[n,4] * 2
        expr = expr - v_flow[N_Q_line + n] * Para.Line[n,5] * 2
        expr = expr + v_flow[N_I_line + n] * Para.Line[n,4] ** 2 
        expr = expr + v_flow[N_I_line + n] * Para.Line[n,5] ** 2
        model.addConstr(expr >= -Para.Big_M * (1 - y_line[n]))
        model.addConstr(expr <=  Para.Big_M * (1 - y_line[n]))
    
    # 3. Second order conic constraint
    for n in range(Para.N_line):
        ep_0 = v_flow[N_P_line + n] * 2
        ep_1 = v_flow[N_Q_line + n] * 2
        ep_2 = v_flow[N_I_line + n] - v_flow[N_V_bus + Para.Line[n,1]]
        ep_3 = v_flow[N_I_line + n] + v_flow[N_V_bus + Para.Line[n,1]]
        model.addConstr(ep_0 * ep_0 + ep_1 * ep_1 + ep_2 * ep_2 <= ep_3 * ep_3)
    
    # 4. Renewables generation
    for n in range(Para.N_gen):
        expr = LinExpr()
        expr = expr + v_flow[N_S_gen + n]
        expr = expr + v_flow[N_C_gen + n]
        G_type = int(Para.Gen[n,4])
        model.addConstr(expr == Para.Gen[n,2] * Para.Day[hour, G_type + 2])
    
    # 5. Lower and Upper bound
    # 1) voltage amplitutde
    for n in range(Para.N_bus):
        model.addConstr(v_flow[N_V_bus  + n] >= Para.V_min)
        model.addConstr(v_flow[N_V_bus  + n] <= Para.V_max)
    # 2) line current
    for n in range(Para.N_line):
        model.addConstr(v_flow[N_I_line + n] >=  0)
        model.addConstr(v_flow[N_I_line + n] <=  y_line[n] * Para.Big_M)
        model.addConstr(v_flow[N_I_line + n] <=  (Para.Line[n,3] / Para.Base_V) ** 2)
    # 3) line flow
    for n in range(Para.N_line):
        smax = Para.Line[n,3]
        # active power
        model.addConstr(v_flow[N_P_line + n] >= -smax)
        model.addConstr(v_flow[N_P_line + n] <=  smax)
        model.addConstr(v_flow[N_P_line + n] >= -y_line[n] * Para.Big_M)
        model.addConstr(v_flow[N_P_line + n] <=  y_line[n] * Para.Big_M)
        # reactive power
        model.addConstr(v_flow[N_Q_line + n] >= -smax)
        model.addConstr(v_flow[N_Q_line + n] <=  smax)
        model.addConstr(v_flow[N_Q_line + n] >= -y_line[n] * Para.Big_M)
        model.addConstr(v_flow[N_Q_line + n] <=  y_line[n] * Para.Big_M)
    # 4) substation
    for n in range(Para.N_sub):
        smax = Para.Sub[n,2]
        model.addConstr(v_flow[N_P_sub  + n] >=  0)
        model.addConstr(v_flow[N_P_sub  + n] <=  smax)
        model.addConstr(v_flow[N_Q_sub  + n] >=  0)
        model.addConstr(v_flow[N_Q_sub  + n] <=  smax)
    # 5) renewables
    for n in range(Para.N_gen):
        G_type = int(Para.Gen[n,4])
        smax = Para.Gen[n,2] * Para.Day[hour, G_type + 2]
        model.addConstr(v_flow[N_S_gen  + n] >=  0)
        model.addConstr(v_flow[N_S_gen  + n] <=  smax)
        model.addConstr(v_flow[N_C_gen  + n] >=  0)
        model.addConstr(v_flow[N_C_gen  + n] <=  smax)
    # 6) load shedding
    for n in range(Para.N_bus):
        smax = Para.Bus[n,1]
        model.addConstr(v_flow[N_P_cut  + n] >=  0)
        model.addConstr(v_flow[N_P_cut  + n] <=  smax)
    for n in range(Para.N_bus):
        smax = Para.Bus[n,2]
        model.addConstr(v_flow[N_Q_cut  + n] >=  0)
        model.addConstr(v_flow[N_Q_cut  + n] <=  smax)

    # Return
    return model


# This function saves results in an Excel file
#
def Save_Result(Para, Result, Name):
    # Save results
    book  = pyxl.Workbook()  # create work-book
    sheet = book.active      # create work-sheet
    sheet['A1'] = "Line"
    sheet['B1'] = "Power"
    for i in range(len(Result)):
        sheet.append(Result[i,:].tolist())  # write data
    book.save(filename = Name)

# This function saves results in an Excel file
#
def Save_Time(Para, Time, Name):
    # Save results
    book  = pyxl.Workbook()  # create work-book
    sheet = book.active      # create work-sheet
    sheet['A1'] = "Hour"
    sheet['B1'] = "IEEE-33"
    sheet['C1'] = "IEEE-69"
    for i in range(len(Time)):
        sheet.append(Time[i,:].tolist())  # write data
    book.save(filename = Name)


if __name__ == "__main__":

    mode = 2

    if mode == 0:
        Name = "data/Data-IEEE-33.xlsx"  # file name
        Para = Parameter(Name)    # System parameter
        Ret0 = Global_Index(Para)
        hour = 12
        t0   = time.perf_counter()
        sol  = Contingency_Screening(Para, hour)
        t1   = time.perf_counter()
        print("Time used: %.2f" % (t1-t0), "s")
        Save_Result(Para, sol, "result/result_traverse_33.xlsx")
    
    if mode == 1:
        Name = "data/Data-IEEE-69.xlsx"  # file name
        Para = Parameter(Name)    # System parameter
        Ret0 = Global_Index(Para)
        hour = 19
        t0   = time.perf_counter()
        sol  = Contingency_Screening(Para, hour)
        t1   = time.perf_counter()
        print("Time used: %.2f" % (t1-t0), "s")
        Save_Result(Para, sol, "result/result_traverse_69.xlsx")

    if mode == 2:
        N_hour = 24
        time_array = np.zeros((N_hour,3))
        for h in range(N_hour):
            time_array[h,0] = h

        # Input parameter
        Name = "data/Data-IEEE-33.xlsx"  # file name
        Para = Parameter(Name)    # System parameter
        Ret0 = Global_Index(Para)

        # Optimizing

        for hour in range(N_hour):
            t0   = time.perf_counter()
            sol  = Contingency_Screening(Para, hour)
            t1   = time.perf_counter()
            print("Time used: %.2f" % (t1-t0), "s")
            time_array[hour,1] = t1-t0

        # Input parameter
        Name = "data/Data-IEEE-69.xlsx"  # file name
        Para = Parameter(Name)    # System parameter
        Ret0 = Global_Index(Para)

        # Optimizing
        for hour in range(N_hour):
            t0   = time.perf_counter()
            sol  = Contingency_Screening(Para, hour)
            t1   = time.perf_counter()
            print("Time used: %.2f" % (t1-t0), "s")
            time_array[hour,2] = t1-t0
        
        Save_Time(Para, time_array, "result/result_traverse_time.xlsx")